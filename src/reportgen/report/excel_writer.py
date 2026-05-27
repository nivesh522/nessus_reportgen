from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..models import Finding
from .themes import Theme


class ExcelReportWriter:
    def __init__(self, theme: Theme):
        self.theme = theme
        self.wb = Workbook()
        self.header_font = Font(bold=True, color=theme.header_font)
        self.header_fill = PatternFill(
            start_color=theme.header_bg, end_color=theme.header_bg, fill_type="solid"
        )
        self.cell_font = Font(color=theme.cell_font)
        self.cell_fill = PatternFill(
            start_color=theme.cell_bg, end_color=theme.cell_bg, fill_type="solid"
        )
        self.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def write(self, output_path: Path, findings: list[Finding]) -> None:
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        self._create_executive_summary(findings)
        self._create_host_summary(findings)
        self._create_findings_sheet(findings)
        self._create_severity_sheets(findings)

        self.wb.save(str(output_path))
        logger.info(f"Report written to {output_path}")

    def _create_executive_summary(self, findings: list[Finding]) -> None:
        ws = self.wb.create_sheet("Executive Summary")
        ws.title = "Executive Summary"

        severity_counts: dict[str, int] = {
            "Critical": 0,
            "High": 0,
            "Medium": 0,
            "Low": 0,
            "Info": 0,
        }

        for f in findings:
            sev = f.severity.capitalize()
            if sev in severity_counts:
                severity_counts[sev] += 1
            else:
                severity_counts["Info"] += 1

        headers = ["Severity", "Count"]
        ws.append(headers)
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.alignment
            cell.border = self.thin_border

        row_num = 2
        for sev in ["Critical", "High", "Medium", "Low", "Info"]:
            count = severity_counts[sev]
            ws.append([sev, count])
            cell_sev = ws.cell(row=row_num, column=1)
            cell_count = ws.cell(row=row_num, column=2)
            for cell in [cell_sev, cell_count]:
                cell.font = self.cell_font
                cell.fill = self.cell_fill
                cell.alignment = self.alignment
                cell.border = self.thin_border
            bg_map = {
                "Critical": self.theme.critical_bg,
                "High": self.theme.high_bg,
                "Medium": self.theme.medium_bg,
                "Low": self.theme.low_bg,
                "Info": self.theme.info_bg,
            }
            cell_sev.fill = PatternFill(
                start_color=bg_map[sev], end_color=bg_map[sev], fill_type="solid"
            )
            row_num += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.5
            ws.column_dimensions[column].width = adjusted_width

    def _create_host_summary(self, findings: list[Finding]) -> None:
        ws = self.wb.create_sheet("Host Summary")

        host_data: dict[str, dict[str, int]] = {}
        for f in findings:
            for ip in f.affected_ips:
                if ip not in host_data:
                    host_data[ip] = {
                        "Total": 0,
                        "Critical": 0,
                        "High": 0,
                        "Medium": 0,
                        "Low": 0,
                        "Info": 0,
                    }
                host_data[ip]["Total"] += 1
                sev = f.severity.capitalize()
                if sev in host_data[ip]:
                    host_data[ip][sev] += 1
                else:
                    host_data[ip]["Info"] += 1

        headers = ["Host", "Findings", "Critical", "High", "Medium", "Low"]
        ws.append(headers)
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.alignment
            cell.border = self.thin_border

        row_num = 2
        for host, data in sorted(host_data.items()):
            ws.append(
                [host, data["Total"], data["Critical"], data["High"], data["Medium"], data["Low"]]
            )
            for col_num in range(1, 7):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = self.cell_font
                cell.fill = self.cell_fill
                cell.alignment = self.alignment
                cell.border = self.thin_border
            row_num += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.5
            ws.column_dimensions[column].width = adjusted_width

    def _create_findings_sheet(self, findings: list[Finding]) -> None:
        ws = self.wb.create_sheet("Findings")

        row_num = 1
        for finding in findings:
            fields = [
                ("Synopsis", finding.synopsis or "N/A"),
                ("Ports", ", ".join(finding.ports) if finding.ports else "N/A"),
                ("Severity", finding.severity),
                ("Risk", finding.risk or "N/A"),
                ("Description", finding.description or "N/A"),
                ("Solution", finding.solution or "N/A"),
                ("CVE numbers", ", ".join(finding.cves) if finding.cves else "N/A"),
                (
                    "Affected IP addresses",
                    ", ".join(finding.affected_ips) if finding.affected_ips else "N/A",
                ),
            ]

            for label, value in fields:
                cell_label = ws.cell(row=row_num, column=1)
                cell_value = ws.cell(row=row_num, column=2)
                cell_label.value = label
                cell_label.font = self.header_font
                cell_label.fill = self.header_fill
                cell_label.alignment = self.alignment
                cell_label.border = self.thin_border

                cell_value.value = value
                cell_value.font = self.cell_font
                cell_value.fill = self.cell_fill
                cell_value.alignment = self.alignment
                cell_value.border = self.thin_border

                row_num += 1
            row_num += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 80

    def _create_severity_sheets(self, findings: list[Finding]) -> None:
        severities = ["Critical", "High", "Medium", "Low", "Info"]
        for sev in severities:
            ws = self.wb.create_sheet(sev)
            sev_findings = [f for f in findings if f.severity.lower() == sev.lower()]
            if not sev_findings:
                ws.append([f"No {sev} findings"])
                continue

            row_num = 1
            for finding in sev_findings:
                fields = [
                    ("Synopsis", finding.synopsis or "N/A"),
                    ("Ports", ", ".join(finding.ports) if finding.ports else "N/A"),
                    ("Severity", finding.severity),
                    ("Risk", finding.risk or "N/A"),
                    ("Description", finding.description or "N/A"),
                    ("Solution", finding.solution or "N/A"),
                    ("CVE numbers", ", ".join(finding.cves) if finding.cves else "N/A"),
                    (
                        "Affected IP addresses",
                        ", ".join(finding.affected_ips) if finding.affected_ips else "N/A",
                    ),
                ]

                for label, value in fields:
                    cell_label = ws.cell(row=row_num, column=1)
                    cell_value = ws.cell(row=row_num, column=2)
                    cell_label.value = label
                    cell_label.font = self.header_font
                    cell_label.fill = self.header_fill
                    cell_label.alignment = self.alignment
                    cell_label.border = self.thin_border

                    cell_value.value = value
                    cell_value.font = self.cell_font
                    cell_value.fill = self.cell_fill
                    cell_value.alignment = self.alignment
                    cell_value.border = self.thin_border

                    row_num += 1
                row_num += 1

            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 80
